"""
Generates two Survey123-ready XLSForm (.xlsx) files:
  1. richland_community_voice.xlsx  — branding feedback survey
  2. richland_stay_engaged.xlsx     — contact / engagement sign-up

Run:  pip install openpyxl
      python generate_xlsforms.py

Then open each file in Survey123 Connect (File → Open XLSForm) to publish.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FOREST = "2B332D"
IVORY  = "F7F5F0"
SAND   = "E8E2D6"
SAGE   = "6B7F5B"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def write_sheet(wb, sheet_name, headers, rows):
    ws = wb.create_sheet(sheet_name)

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=IVORY, size=10)
        cell.fill = PatternFill(start_color=FOREST, end_color=FOREST, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows — alternate light fill
    for r, row in enumerate(rows, 2):
        fill_color = SAND if r % 2 == 0 else IVORY
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Auto-width (capped at 60)
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 2, 60)

    ws.row_dimensions[1].height = 18
    return ws


def make_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet
    return wb


def save(wb, path):
    wb.save(path)
    print(f"  Saved -> {path}")


# ---------------------------------------------------------------------------
# Survey 1: Community Voice
# ---------------------------------------------------------------------------

def build_community_voice():
    wb = make_workbook()

    # ── settings ──────────────────────────────────────────────────────────
    write_sheet(wb, "settings",
        ["form_title", "form_id", "version", "default_language"],
        [["Richland Parish — Community Voice", "richland_community_voice", "2026070801", "English"]]
    )

    # ── choices ───────────────────────────────────────────────────────────
    choices_headers = ["list_name", "name", "label"]
    choices_rows = [
        # how_long
        ["how_long", "less_1",       "Less than 1 year"],
        ["how_long", "1_5",          "1–5 years"],
        ["how_long", "5_10",         "5–10 years"],
        ["how_long", "10_20",        "10–20 years"],
        ["how_long", "20_plus",      "20+ years"],
        ["how_long", "not_resident", "I don't live here but I'm connected to the parish"],

        # parish_area
        ["parish_area", "rayville",     "Rayville area"],
        ["parish_area", "delhi",        "Delhi area"],
        ["parish_area", "start",        "Start area"],
        ["parish_area", "mangham",      "Mangham area"],
        ["parish_area", "other_area",   "Other area"],
        ["parish_area", "parish_wide",  "Throughout the parish"],

        # words — word association
        ["words", "agricultural",  "Agricultural"],
        ["words", "close_knit",    "Close-knit"],
        ["words", "family",        "Family-oriented"],
        ["words", "hardworking",   "Hardworking"],
        ["words", "historic",      "Historic"],
        ["words", "natural",       "Natural beauty"],
        ["words", "quiet",         "Quiet"],
        ["words", "resilient",     "Resilient"],
        ["words", "rural",         "Rural"],
        ["words", "southern",      "Southern"],
        ["words", "traditional",   "Traditional"],
        ["words", "welcoming",     "Welcoming"],
        ["words", "growing",       "Growing"],
        ["words", "proud",         "Proud"],
        ["words", "authentic",     "Authentic"],

        # strengths
        ["strengths", "agriculture",     "Agriculture & farming heritage"],
        ["strengths", "community",       "Strong community ties"],
        ["strengths", "affordability",   "Affordability"],
        ["strengths", "safety",          "Safety"],
        ["strengths", "schools",         "Schools & education"],
        ["strengths", "nature",          "Natural resources & outdoors"],
        ["strengths", "history",         "History & culture"],
        ["strengths", "location",        "Location & accessibility"],
        ["strengths", "hunting_fishing", "Hunting & fishing"],
        ["strengths", "faith",           "Faith community"],

        # priorities
        ["priorities", "economic",     "Economic development & jobs"],
        ["priorities", "housing",      "Housing options"],
        ["priorities", "roads",        "Roads & infrastructure"],
        ["priorities", "parks",        "Parks & recreation"],
        ["priorities", "downtown",     "Downtown revitalization"],
        ["priorities", "broadband",    "Broadband & connectivity"],
        ["priorities", "healthcare",   "Healthcare access"],
        ["priorities", "agriculture",  "Supporting agriculture"],
        ["priorities", "youth",        "Opportunities for youth"],
        ["priorities", "environment",  "Environmental stewardship"],
    ]
    write_sheet(wb, "choices", choices_headers, choices_rows)

    # ── survey ────────────────────────────────────────────────────────────
    survey_headers = [
        "type", "name", "label", "hint",
        "required", "required_message",
        "relevant", "appearance", "parameters"
    ]
    survey_rows = [
        # Welcome
        ["note", "welcome",
         "Thank you for taking a few minutes to share your voice.\n\nYour answers will directly shape the identity and direction of the Richland Parish Master Plan.",
         "", "", "", "", "", ""],

        # ── Group: About You ──
        ["begin group", "about_you", "About You", "", "", "", "", "field-list", ""],

        ["select_one how_long", "years_in_parish",
         "How long have you lived in or been connected to Richland Parish?",
         "", "yes", "Please answer this question.", "", "", ""],

        ["select_one parish_area", "parish_area",
         "Which area of Richland Parish do you primarily live or work in?",
         "", "no", "", "", "", ""],

        ["end group", "about_you", "", "", "", "", "", "", ""],

        # ── Group: Your Voice ──
        ["begin group", "your_voice", "Your Voice", "", "", "", "", "", ""],

        ["select_multiple words", "word_association",
         "Which words best describe Richland Parish to you?",
         "Choose up to 5 words that feel right.", "no", "", "", "", "randomize=true"],

        ["text", "word_other",
         "Any other words you'd use to describe Richland Parish?",
         "Feel free to write in your own.", "no", "", "", "multiline", ""],

        ["text", "what_it_means",
         "In your own words, what does Richland Parish mean to you?",
         "There are no wrong answers — share whatever comes to mind.",
         "yes", "We'd love to hear your thoughts.", "", "multiline", ""],

        ["text", "preserve",
         "What's one thing about Richland Parish you most want to preserve?",
         "", "no", "", "", "multiline", ""],

        ["text", "improve",
         "What's one thing you'd most like to see change or improve?",
         "", "no", "", "", "multiline", ""],

        ["select_multiple strengths", "top_strengths",
         "What are Richland Parish's greatest strengths?",
         "Choose up to 3.", "no", "", "", "", ""],

        ["select_multiple priorities", "future_priorities",
         "What should the Master Plan prioritize?",
         "Choose up to 3 topics that matter most to you.", "no", "", "", "", ""],

        ["end group", "your_voice", "", "", "", "", "", "", ""],

        # ── Group: Location (optional) ──
        ["begin group", "location_group", "Your Location (Optional)", "", "", "", "", "", ""],

        ["note", "location_note",
         "The next question is optional. Your location helps us understand where responses are coming from across the parish.",
         "", "", "", "", "", ""],

        ["geopoint", "response_location",
         "Where in Richland Parish are you?",
         "Tap to place a pin at your approximate location. This is completely optional.",
         "no", "", "", "", ""],

        ["end group", "location_group", "", "", "", "", "", "", ""],

        # Final
        ["text", "additional_comments",
         "Any other thoughts you'd like to share with the planning team?",
         "", "no", "", "", "multiline", ""],

        ["note", "thank_you",
         "Thank you for participating! Your response has been recorded.\n\nStay connected — visit the project page to sign up for updates.",
         "", "", "", "", "", ""],
    ]
    write_sheet(wb, "survey", survey_headers, survey_rows)

    save(wb, "richland_community_voice.xlsx")


# ---------------------------------------------------------------------------
# Survey 2: Stay Engaged
# ---------------------------------------------------------------------------

def build_stay_engaged():
    wb = make_workbook()

    # ── settings ──────────────────────────────────────────────────────────
    write_sheet(wb, "settings",
        ["form_title", "form_id", "version", "default_language"],
        [["Richland Parish — Stay Engaged", "richland_stay_engaged", "2026070801", "English"]]
    )

    # ── choices ───────────────────────────────────────────────────────────
    choices_rows = [
        # contact_pref
        ["contact_pref", "email",  "Email"],
        ["contact_pref", "text",   "Text message"],
        ["contact_pref", "either", "Either is fine"],

        # interests
        ["interests", "economic",   "Economic development & jobs"],
        ["interests", "housing",    "Housing"],
        ["interests", "roads",      "Roads & transportation"],
        ["interests", "parks",      "Parks & recreation"],
        ["interests", "zoning",     "Land use & zoning"],
        ["interests", "agriculture","Agriculture"],
        ["interests", "downtown",   "Downtown development"],
        ["interests", "youth",      "Youth & families"],
        ["interests", "environment","Environment & natural resources"],
        ["interests", "historic",   "Historic preservation"],
        ["interests", "all",        "All of the above"],

        # how_heard
        ["how_heard", "social_media",  "Social media"],
        ["how_heard", "friend_family", "Friend or family member"],
        ["how_heard", "newspaper",     "Newspaper or news outlet"],
        ["how_heard", "flyer",         "Flyer or poster"],
        ["how_heard", "email_list",    "Email or newsletter"],
        ["how_heard", "parish_office", "Parish or city office"],
        ["how_heard", "event",         "Community event"],
        ["how_heard", "other",         "Other"],

        # attend_meetings
        ["attend_meetings", "yes",   "Yes, I'd like to attend in person"],
        ["attend_meetings", "virtual","Yes, but I prefer virtual options"],
        ["attend_meetings", "either","Either works for me"],
        ["attend_meetings", "no",    "No, but please keep me updated online"],
    ]
    write_sheet(wb, "choices", ["list_name", "name", "label"], choices_rows)

    # ── survey ────────────────────────────────────────────────────────────
    survey_headers = [
        "type", "name", "label", "hint",
        "required", "required_message", "relevant", "appearance", "constraint", "constraint_message"
    ]
    survey_rows = [
        ["note", "welcome",
         "Stay connected with the Richland Parish Master Plan process.\n\nSign up to receive updates, meeting invitations, and ways to participate as the plan develops.",
         "", "", "", "", "", "", ""],

        # ── Group: Contact Info ──
        ["begin group", "contact_info", "Your Contact Information", "", "", "", "", "field-list", "", ""],

        ["text", "first_name", "First Name", "", "yes", "Please enter your first name.", "", "", "", ""],
        ["text", "last_name",  "Last Name",  "", "yes", "Please enter your last name.",  "", "", "", ""],

        ["text", "email",
         "Email Address",
         "", "yes", "Please enter a valid email address.", "",
         "email", "regex(., '^[a-zA-Z0-9._%+\\-]+@[a-zA-Z0-9.\\-]+\\.[a-zA-Z]{2,}$')",
         "Please enter a valid email address (e.g. name@example.com)."],

        ["text", "phone",
         "Phone Number",
         "Optional — for text message updates.", "no", "", "", "tel", "", ""],

        ["select_one contact_pref", "contact_preference",
         "How do you prefer to be contacted?",
         "", "no", "", "", "", "", ""],

        ["end group", "contact_info", "", "", "", "", "", "", "", ""],

        # ── Group: Your Interests ──
        ["begin group", "interests_group", "Your Interests", "", "", "", "", "", "", ""],

        ["select_multiple interests", "areas_of_interest",
         "What topics are you most interested in?",
         "Select all that apply.", "no", "", "", "", "", ""],

        ["select_one attend_meetings", "meeting_interest",
         "Would you like to attend public meetings or workshops?",
         "", "no", "", "", "", "", ""],

        ["text", "meeting_availability",
         "What times work best for you?",
         "e.g., weekday evenings, Saturday mornings, lunch hours",
         "no", "", "${meeting_interest}!='no'", "multiline", "", ""],

        ["select_one how_heard", "how_heard",
         "How did you hear about this project?",
         "", "no", "", "", "", "", ""],

        ["end group", "interests_group", "", "", "", "", "", "", "", ""],

        ["text", "additional_comments",
         "Any questions or comments for the planning team?",
         "", "no", "", "", "multiline", "", ""],

        ["note", "thank_you",
         "You're signed up! We'll be in touch with updates about the Richland Parish Master Plan.\n\nThank you for being part of this process.",
         "", "", "", "", "", "", ""],
    ]
    write_sheet(wb, "survey", survey_headers, survey_rows)

    save(wb, "richland_stay_engaged.xlsx")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Generating Survey123 XLSForms...")
    build_community_voice()
    build_stay_engaged()
    print("\nDone. Next steps:")
    print("  1. Open Survey123 Connect")
    print("  2. File > Open XLSForm > select each .xlsx file")
    print("  3. Review questions, then click Publish")
    print("  4. Copy each survey's share URL into src/config/surveys.js")
